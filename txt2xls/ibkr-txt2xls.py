#!/usr/bin/env python3

import argparse
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


GROUP_BREAK_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFD9D9D9",
)
TWO_DECIMAL_FORMAT = "0.00"

INTEGER_RE = re.compile(r"^[+-]?\d+$")
DECIMAL_RE = re.compile(r"^[+-]?(?:\d+\.\d*|\.\d+)$")


def cell_value(token):
    token = token.replace("=", "")
    if INTEGER_RE.fullmatch(token):
        return int(token)
    if DECIMAL_RE.fullmatch(token):
        return float(token)
    return token


def is_group_break(line):
    return "Bestand" in line or re.search(r"={2,}", line) is not None


def column_slices(header_line):
    # The source report right-aligns headings in their fixed-width columns,
    # so each heading's end position marks that column's end.
    headings = list(re.finditer(r"\S+", header_line))
    if not headings:
        raise ValueError("first line does not contain any column headings")

    ends = []
    index = 0
    while index < len(headings):
        heading = headings[index]
        next_heading = headings[index + 1] if index + 1 < len(headings) else None

        if (
            next_heading
            and next_heading.start() == heading.end() + 1
            and (
                re.fullmatch(r"\d{4}-\d{2}", heading.group())
                and next_heading.group() == "DD"
            )
        ):
            ends.append(next_heading.end())
            index += 2
        else:
            ends.append(heading.end())
            index += 1

    starts = [0] + ends[:-1]
    return list(zip(starts, ends[:-1] + [None]))


def parse_fixed_width(line, slices):
    return [line[start:end].replace("=", "").strip() for start, end in slices]


def convert(input_path, output_path=None):
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.with_suffix(".xlsx")
    else:
        output_path = Path(output_path)

    workbook = Workbook()
    sheet = workbook.active
    #sheet.title = "Output"

    with input_path.open("r", encoding="utf-8") as input_file:
        lines = [line.rstrip("\n\r") for line in input_file]

    if not lines:
        raise ValueError(f"{input_path} is empty")

    slices = column_slices(lines[0])
    header_tokens = parse_fixed_width(lines[0], slices)
    two_decimal_columns = {
        column_index
        for column_index, token in enumerate(header_tokens, start=1)
        if token.endswith(")")
    }
    max_widths = [0] * len(slices)

    for row_index, line in enumerate(lines, start=1):
        tokens = parse_fixed_width(line, slices)
        grey_row = is_group_break(line)

        for column_index, token in enumerate(tokens, start=1):
            if not token:
                if grey_row:
                    cell = sheet.cell(row=row_index, column=column_index, value="")
                    cell.fill = GROUP_BREAK_FILL
                continue

            value = cell_value(token)
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if column_index in two_decimal_columns and isinstance(value, (int, float)):
                cell.number_format = TWO_DECIMAL_FORMAT
            if grey_row:
                cell.fill = GROUP_BREAK_FILL

            max_widths[column_index - 1] = max(
                max_widths[column_index - 1],
                len(str(token)),
            )

    for column_index, width in enumerate(max_widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(width + 2, 8),
            30,
        )

    workbook.save(output_path)
    return output_path


def parse_args(argv):
    parser = argparse.ArgumentParser(
        description="Convert a fixed-width text file to XLSX.",
    )
    parser.add_argument("filename", help="text file to convert")
    parser.add_argument(
        "-o",
        "--output",
        help="output XLSX filename, defaults to input filename with .xlsx suffix",
    )
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(sys.argv[1:] if argv is None else argv)
    output_path = convert(args.filename, args.output)
    print(output_path)


if __name__ == "__main__":
    main()
